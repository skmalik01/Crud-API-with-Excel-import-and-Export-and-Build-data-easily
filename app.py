from flask import Flask, jsonify, request, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
import openpyxl

app = Flask(__name__)

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///students.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = True

db = SQLAlchemy(app)
ma = Marshmallow(app)

class Students(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(40))
    age = db.Column(db.Integer)
    gender = db.Column(db.String(10))
    email = db.Column(db.String(40))
    course = db.Column(db.String(20))
    grade = db.Column(db.String(10))
    city = db.Column(db.String(20))
    
class StudentsSchema(ma.Schema):
    class Meta:
        model = Students
        fields = ["id", "name", "age", "gender", "email", "course", "grade", "city"]
    
student_schema = StudentsSchema()
students_schema = StudentsSchema(many=True)

with app.app_context():
    db.create_all()
    
@app.route("/students", methods=["POST"])
def add_students():
    data = request.get_json()
    new_data = student_schema.load(data)
    student = Students(**new_data)
    db.session.add(student)
    db.session.commit()
    return jsonify({"message" : "Successfully Added"})

@app.route("/students", methods=["GET"])
def get_all_student():
    all_student = Students.query.all()
    return students_schema.dump(all_student)

@app.route("/students/<int:student_id>", methods=["GET"])
def get_student(student_id):
    student = Students.query.get(student_id)
    if student:
        return student_schema.dump(student)
    return jsonify({"message" : "Student not found"})

@app.route("/students/import", methods=["POST"])
def import_excel():
    files = request.files["filename"]
    wb = openpyxl.load_workbook(files)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        new_student = Students(name=row[0], age=row[1], gender=row[2], email=row[3], course=row[4], grade=row[5], city=row[6])
        db.session.add(new_student)
    db.session.commit()
    return jsonify({"message" : "File Imported"})
    
@app.route("/students/export", methods=["GET"])
def export_excel():
    students = Students.query.all()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Name", "Age", "Gender", "Email", "Course", "Grade", "City"])
    for student in students:
        sheet.append([student.name, student.age, student.gender, student.email, student.course, student.grade, student.city])
    filename = 'Student.xlsx'
    wb.save("filename")
    return send_file(filename, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True, port=5050)